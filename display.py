import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import keras.backend as K
import tensorflow as tf
import openpyxl
import os


def plot_turbine_ids(data, tag, folder='plots', maxbars=-1, series=True):
    """
    Plot turbine ID distribution with sorted ID number in x axes and number of ID occurrences in y axes
    :param data: turbine IDs data. It can be a series (series = True) or an ID counter (series = False)
    :param tag: plot reference name
    :param folder: folder to store the plots
    :param maxbars: maximum number of IDs to show (-1 to show all IDs in the input data)
    :param series: Indicate if the input data is a pandas series (series = True) or an ID counter (series = False)
    :return: None (it saves a png file with the plot)
    """
    fig, ax = plt.subplots(figsize=(20, 10))
    plt.rcParams['font.size'] = 14
    data = pd.Series(data)
    data = data.astype(int)

    if series:
        if maxbars == -1:
            maxbars = data.value_counts().shape[0]
        datapoints = data.value_counts().sum()
        data.value_counts().sort_index()[:maxbars].plot(kind='bar', ax=ax)
    else:
        if maxbars == -1:
            maxbars = data.shape[0]
        datapoints = data.to_numpy().sum()
        data[:maxbars].plot(kind='bar', ax=ax)
    ax.set_ylabel('NUMBER OF OCCURRENCES', fontsize=16, fontweight='bold')
    ax.set_xlabel('ID MESSAGES', fontsize=16, fontweight='bold')
    ax.set_title(tag.upper() + ' (' + str(datapoints) + ' datapoints)', fontweight='bold', fontsize=24)
    ax.tick_params(axis='x', labelrotation=70)
    ax.tick_params(axis='both', labelsize=14)
    ax.bar_label(ax.containers[0])
    ax.grid(visible=True)
    fig.tight_layout()
    plt.savefig(os.path.join(os.getcwd(), folder, tag + '.png'), bbox_inches='tight')
    plt.close()


def plot_results(data, test, met, loss, tag=''):
    """
    Create a plot with the train and validation information during simulation per each epoch for model loss and metrics
    :param data: history for simulated model
    :param test: test set evaluation results
    :param met: list of metrics name for plot references
    :param loss: loss name for plot references
    :param tag: plot reference name
    :return: train and validation data for best epoch (it saves a png file with the plot)
    """
    if tag != '':
        tag = ' ' + tag
    fig, axes = plt.subplots(nrows=-(-len(met)//2), ncols=2, figsize=(20, 10))
    ax = axes.ravel()
    best_ind = np.argmin(data['val_loss'])
    train_data = np.empty([len(met)])
    val_data = np.empty([len(met)])
    for i in range(len(met)):
        if i > 0:
            loss = met[i]
        ax[i].plot(range(1, len(data[met[i]]) + 1), data[met[i]], ls='-', lw=2, color='b',
                   label='Train ' + loss)
        ax[i].plot(range(1, len(data[met[i]]) + 1), data['val_' + met[i]], ls='--', lw=2, color='b',
                   label='Validation ' + loss)
        ax[i].set_xlabel('EPOCHS', fontweight='bold', fontsize=12)
        ax[i].set_ylabel(met[i].upper(), fontweight='bold', fontsize=12)
        text_str = '\nTRAIN / VAL / TEST ' + loss.upper() + ' = {:.4f} / {:.4f} / {:.4f}'.format(
            data[met[i]][best_ind], data['val_' + met[i]][best_ind], test[i])
        train_data[i] = data[met[i]][best_ind]
        val_data[i] = data['val_' + met[i]][best_ind]
        ax[i].set_title(text_str, fontweight='bold', fontsize=12)
        ax[i].grid(visible=True)
        ax[i].legend()
    text_str = '\nBEST MODEL SAVED AT EPOCH {}'.format(best_ind + 1)
    plt.suptitle('SIMULATION RESULTS' + tag.upper() + text_str, fontweight='bold', fontsize=16)
    fig.tight_layout()
    plt.savefig(os.path.join(os.getcwd(), 'plots', 'Simulation results' + tag + '.png'), bbox_inches='tight')
    plt.close()
    return train_data, val_data


def plot_confusion_matrix(trues, preds, tag=''):
    """
    Create confusion matrix for single label output simulation and generate metrics calculations as accuracy, precision
    and F1 score per each output class and global averages
    :param trues: real target
    :param preds: model predicted target
    :param tag: plot reference name
    :return: accuracy, precision and F1 score per each output class (it saves a png file with the plot)
    """
    if tag != '':
        tag = ' ' + tag
    if len(preds.shape) == 2:
        seq = False
    else:
        seq = True
    n = K.int_shape(preds)[-1]
    y_pred = np.argmax(preds, axis=-1)
    y_true = np.argmax(trues, axis=-1)
    y_true = y_true.astype(int)
    confusion_matrix = np.zeros([n, n])
    if seq:
        for i in range(y_pred.shape[0]):
            for j in range(y_pred.shape[1]):
                confusion_matrix[y_pred[i, j], y_true[i, j]] += 1
    else:
        for i in range(y_pred.shape[0]):
            confusion_matrix[y_pred[i], y_true[i]] += 1
    conf_matrix = confusion_matrix.astype(int)

    fig, ax = plt.subplots(figsize=(20, 10))
    plt.pcolormesh(conf_matrix, cmap=plt.get_cmap('cool'))  # plt.colorbar()
    labels = np.arange(conf_matrix.shape[0])
    yrange = [x + 0.5 for x in range(conf_matrix.shape[0])]
    xrange = [x + 0.5 for x in range(conf_matrix.shape[1])]
    plt.xticks(xrange, labels, rotation=75, ha='center')
    ax.xaxis.tick_top()
    plt.yticks(yrange, labels, va='center')
    for i in range(len(xrange)):
        for j in range(len(yrange)):
            ax.text(xrange[i], yrange[j], str(conf_matrix[j, i]),
                    ha="center", va="center", color="k", fontsize=10)
    plt.xlabel("TRUE ID", weight='bold', fontsize=16)
    plt.ylabel("PREDICTION ID", weight='bold', fontsize=16)

    # Total accuracy
    ok = np.trace(conf_matrix)
    tot = np.sum(conf_matrix)
    acc_tot = round(100 * ok / tot, 1)

    # Detailed accuracy/precision
    acc = np.empty([conf_matrix.shape[0]])
    prec = np.empty([conf_matrix.shape[0]])
    f1 = np.empty([conf_matrix.shape[0]])
    for i in range(conf_matrix.shape[0]):
        acc[i] = min(100, max(0, 100 * conf_matrix[i, i] / (np.sum(conf_matrix[:, i]) + K.epsilon())))
        prec[i] = min(100, max(0, 100 * conf_matrix[i, i] / (np.sum(conf_matrix[i, :]) + K.epsilon())))
        f1[i] = 2 / (1 / (prec[i] + K.epsilon()) + 1 / (acc[i] + K.epsilon()))
    acc_macro = np.round(np.mean(acc), 1)
    prec_macro = np.round(np.mean(prec), 1)
    f1_macro = np.round(np.mean(f1), 1)

    t1_str = 'TEST ID ACC:'
    t2_str = 'TEST ID PREC:'
    t3_str = 'TEST ID F1:'
    for i in range(acc.shape[0]):
        if i == round(acc.shape[0] / 2):
            n = '\n'
        else:
            n = ''
        t1_str += n + ' ' + str(i) + '=' + str(round(acc[i], 1)) + '%'
        t2_str += n + ' ' + str(i) + '=' + str(round(prec[i], 1)) + '%'
        t3_str += n + ' ' + str(i) + '=' + str(round(f1[i], 1)) + '%'

    fig.suptitle('CONFUSION MATRIX' + tag + '\nTotal Accuracy=' + str(acc_tot) + '% Macro Accuracy=' + str(acc_macro)
                 + '% Macro Precision=' + str(prec_macro) + '% Macro F1 score=' + str(f1_macro) + '%', weight='bold',
                 fontsize=14)
    ax.set_title(t1_str + '\n\n' + t2_str + '\n\n' + t3_str, weight='bold', fontsize=12)
    ax.tick_params(axis='both', labelsize=12)
    fig.tight_layout()
    plt.savefig(os.path.join(os.getcwd(), 'plots', 'Confusion Matrix' + tag + '.png'), bbox_inches='tight')
    plt.close()
    return acc, prec, f1


def plot_classification_report(y_true, y_pred, multilabel=False, tag=''):
    """
    Generate classification report in a table format calculating true positives, false positives, false negatives,
    recall, precision, F1 score per each class and global averages. Valid for multilabel applications considering
    positive when probability > 0.5
    :param y_true: real target
    :param y_pred: model predicted target
    :param multilabel: True for multilabel application and False for single label application
    :param tag: plot reference name
    :return: None (it saves a png file with the plot)
    """
    if tag != '':
        tag = ' ' + tag
    if multilabel:
        y_pred = K.round(y_pred)
    else:
        y_pred /= K.max(y_pred, axis=-1, keepdims=True)  # assign 1 to the predicted class
    n = K.int_shape(y_pred)[-1]
    results = np.zeros([n + 2, 7])
    for c in range(n):
        results[c, 0] = c
    ax = [i for i in range(len(K.int_shape(y_pred)) - 1)]
    results[:n, 1] = K.sum(tf.cast(tf.logical_and(K.equal(y_true, 1), K.equal(y_pred, 1)), 'float32'), axis=ax)
    results[:n, 2] = K.sum(tf.cast(tf.logical_and(K.not_equal(y_true, 1), K.equal(y_pred, 1)), 'float32'), axis=ax)
    results[:n, 3] = K.sum(tf.cast(tf.logical_and(K.equal(y_true, 1), K.not_equal(y_pred, 1)), 'float32'), axis=ax)
    results[:n, 4] = 100 * results[:n, 1] / (K.epsilon() + results[:n, 1] + results[:n, 3])
    results[:n, 5] = 100 * results[:n, 1] / (K.epsilon() + results[:n, 1] + results[:n, 2])
    results[:n, 6] = 2 / ((1 / (K.epsilon() + results[:n, 4])) + (1 / (K.epsilon() + results[:n, 5])))
    results[n, 4] = 100 * np.sum(results[:n, 1]) / (K.epsilon() + np.sum(results[:n, 1]) + np.sum(results[:n, 3]))
    results[n + 1, 4] = np.mean(results[:n, 4])
    results[n, 5] = 100 * np.sum(results[:n, 1]) / (K.epsilon() + np.sum(results[:n, 1]) + np.sum(results[:n, 2]))
    results[n + 1, 5] = np.mean(results[:n, 5])
    results[n, 6] = 2 / ((1 / (K.epsilon() + results[n, 4])) + (1 / (K.epsilon() + results[n, 5])))
    results[n + 1, 6] = np.mean(results[:n, 6])
    df = pd.DataFrame(results, columns=['Class', 'TP', 'FP', 'FN', 'TPR %', 'Precision %', 'F1 %'])
    df = df.round(decimals=1)
    df.iloc[n, 0] = 'TOTAL'
    df.iloc[n + 1, 0] = 'MACRO'
    for ind in range(1, 4):
        df.iloc[n, ind] = ' '
        df.iloc[n + 1, ind] = ' '
    df.set_index('Class', inplace=True)
    fig, ax = plt.subplots(figsize=(8, 1))
    t = ax.table(cellText=df.values, rowLabels=df.index, colLabels=df.columns, cellLoc='center')
    t.auto_set_font_size(False)
    t.set_fontsize(10)
    ax.set_title('Classification Report' + tag)
    ax.axis('off')
    ax.axis('tight')
    plt.savefig(os.path.join(os.getcwd(), 'plots', 'Classification Report' + tag + '.png'), bbox_inches='tight')
    plt.close()


def plot_preds_distribution(y_true, y_pred, threshold=0.5, tag=''):
    """
    Generate a special classification report for multilabel applications in which a correct prediction is considered if
    the prediction with the highest probability and > threshold matches with one real label
    :param y_true: real target
    :param y_pred: model predicted target
    :param threshold: minimum probability in pu to consider prediction as True
    :param tag: plot reference name
    :return: None (it saves a png file with the plot)
    """
    if tag != '':
        tag = ' ' + tag
    y_pred_max = K.max(y_pred, axis=-1, keepdims=True)
    y_pred = tf.where(y_pred > threshold, y_pred / y_pred_max, 0)
    n = K.int_shape(y_pred)[-1]
    results = np.zeros([n + 2, 7])
    for c in range(n):
        results[c, 0] = c
    ax = [i for i in range(len(K.int_shape(y_pred)) - 1)]
    results[:n, 1] = K.sum(tf.cast(K.equal(y_true, 1), 'float32'), axis=ax)
    results[n, 1] = np.sum(results[:n, 1])
    results[:n, 2] = K.sum(tf.cast(K.equal(y_pred, 1), 'float32'), axis=ax)
    results[n, 2] = np.sum(results[:n, 2])
    results[:n, 3] = 100 * results[:n, 2] / (K.epsilon() + np.sum(results[:n, 2]))
    results[n, 3] = np.sum(results[:n, 3])
    results[:n, 4] = K.sum(tf.cast(tf.logical_and(K.equal(y_true, 1), K.equal(y_pred, 1)), 'float32'), axis=ax)
    results[n, 4] = np.sum(results[:n, 4])
    results[:n, 5] = K.sum(tf.cast(tf.logical_and(K.equal(y_true, 0), K.equal(y_pred, 1)), 'float32'), axis=ax)
    results[n, 5] = np.sum(results[:n, 5])
    results[:n, 6] = 100 * results[:n, 4] / (K.epsilon() + results[:n, 2])
    results[n, 6] = 100 * np.sum(results[:n, 4]) / (K.epsilon() + np.sum(results[:n, 2]))
    results[n + 1, 6] = np.mean(results[:n, 6])
    df = pd.DataFrame(results, columns=['Class', 'Trues', 'Preds', 'Preds Dist %', 'TP', 'FP', 'Precision %'])
    df = df.round(decimals=1)
    df.iloc[n, 0] = 'TOTAL'
    df.iloc[n + 1, 0] = 'MACRO'
    for ind in range(1, 6):
        df.iloc[n + 1, ind] = ' '
    df.set_index('Class', inplace=True)
    fig, ax = plt.subplots(figsize=(8, 1))
    t = ax.table(cellText=df.values, rowLabels=df.index, colLabels=df.columns, cellLoc='center')
    t.auto_set_font_size(False)
    t.set_fontsize(10)
    rows = K.int_shape(y_pred)[0]
    ok = K.sum(K.clip(K.sum(tf.cast(tf.logical_and(K.equal(y_true, 1), K.equal(y_pred, 1)), 'float32'), axis=-1), 0, 1))
    predicted = K.sum(K.clip(K.sum(tf.cast(K.equal(y_pred, 1), 'float32'), axis=-1), 0, 1))
    ok = tf.cast(ok, 'int32').numpy()
    predicted = tf.cast(predicted, 'int32').numpy()
    sub = 'Total Accuracy = ' + str(round(100*ok/rows, 1)) + '%, Relative Accuracy = ' + \
          str(round(100*ok/predicted, 1)) + '% and Prediction Rate = ' + str(round(100*predicted/rows, 1)) + \
          '% (Total samples ' + str(rows) + ' - Predicted samples ' + str(predicted) + ' - OK samples ' + str(ok) + ')'
    ax.set_title('Prediction Distribution' + tag + '\n\n' + sub)
    ax.axis('off')
    ax.axis('tight')
    plt.savefig(os.path.join(os.getcwd(), 'plots', 'Prediction Distribution' + tag + '.png'), bbox_inches='tight')
    plt.close()


def table_turbine_ids(id_data, tot_ids, folder='plots'):
    """
    Create a table indicating the turbine ID percentage distribution
    :param id_data: turbine IDs data. It can be a series (series = True) or an ID counter (series = False)
    :param tot_ids: plot reference name
    :param folder: folder to store the plots
    :return: None (it saves a png file with the plot)
    """
    id_class = []
    for i in range(tot_ids):
        id_class.append('ID' + str(i))
    columns = ['Turbine Name']
    columns.extend(id_class)
    df = pd.DataFrame(id_data, columns=columns)
    df[id_class] = df[id_class].round(decimals=1)
    df['Turbine Name'] = df['Turbine Name'].round(0).astype(int)
    df.set_index('Turbine Name', inplace=True)
    fig, ax = plt.subplots(figsize=(15, 1))
    t = ax.table(cellText=df.values, rowLabels=df.index, colLabels=df.columns, cellLoc='center')
    t.auto_set_font_size(False)
    t.set_fontsize(10)
    ax.set_title('Turbine ID Failure Percentage Distribution')
    ax.axis('off')
    ax.axis('tight')
    plt.savefig(os.path.join(os.getcwd(), folder, 'Turbine ID failure distribution.png'), bbox_inches='tight')
    plt.close()


def write_results_excel(train, val, test, acc, prec, f1, data, file='Results'):
    """
    Generate an output Excel file summarizing the simulation information as parametrization and sets results
    :param train: train set simulation data in the format --> [loss, metric1, ... , metric_n]
    :param val: validation set simulation data in the format --> [loss, metric1, ... , metric_n]
    :param test: testing set simulation data in the format --> [loss, metric1, ... , metric_n]
    :param acc: accuracy per each output class in the format --> [acc_ID1, acc_ID2, ... , acc_ID_n]
    :param prec: precision per each output class in the format --> [prec_ID1, prec_ID2, ... , prec_ID_n]
    :param f1: F1 score per each output class in the format --> [f1_ID1, f1_ID2, ... , f1_ID_n]
    :param data: simulation parametrization data in the format --> [units1, units2, data_pre, loss_name, multiplier,
    crit_ids, max_samples, dropout, lrate, batch_size, lookback]
    :param file: string for output Excel file name
    :return: None (it saves an Excel file with the simulation data)
    """
    try:
        excel_file = openpyxl.load_workbook(file + '.xlsx')
        sheet = excel_file[excel_file.sheetnames[0]]
    except (Exception,):
        # If Excel file does not exist, create new one
        excel_file = openpyxl.Workbook()
        sheet = excel_file[excel_file.sheetnames[0]]
        sheet.cell(row=1, column=1).value = 'units1'
        sheet.cell(row=2, column=1).value = 'units2'
        sheet.cell(row=3, column=1).value = 'preprocess'
        sheet.cell(row=4, column=1).value = 'loss'
        sheet.cell(row=5, column=1).value = 'multiplier'
        sheet.cell(row=6, column=1).value = 'critical ids'
        sheet.cell(row=7, column=1).value = 'max samples'
        sheet.cell(row=8, column=1).value = 'dropout'
        sheet.cell(row=9, column=1).value = 'learning rate'
        sheet.cell(row=10, column=1).value = 'batch size'
        sheet.cell(row=11, column=1).value = 'lookback'
        sheet.cell(row=12, column=1).value = 'TRAIN loss'
        sheet.cell(row=13, column=1).value = 'VAL loss'
        sheet.cell(row=14, column=1).value = 'TEST loss'
        sheet.cell(row=15, column=1).value = 'TRAIN F1 macro'
        sheet.cell(row=16, column=1).value = 'VAL F1 macro'
        sheet.cell(row=17, column=1).value = 'TEST F1 macro'
        sheet.cell(row=18, column=1).value = 'TRAIN TPR macro'
        sheet.cell(row=19, column=1).value = 'VAL TPR macro'
        sheet.cell(row=20, column=1).value = 'TEST TPR macro'
        sheet.cell(row=21, column=1).value = 'TRAIN TPR total'
        sheet.cell(row=22, column=1).value = 'VAL TPR total'
        sheet.cell(row=23, column=1).value = 'TEST TPR total'
        sheet.cell(row=24, column=1).value = 'TRAIN prec macro'
        sheet.cell(row=25, column=1).value = 'VAL prec macro'
        sheet.cell(row=26, column=1).value = 'TEST prec macro'
        sheet.cell(row=27, column=1).value = 'TRAIN prec total'
        sheet.cell(row=28, column=1).value = 'VAL prec total'
        sheet.cell(row=29, column=1).value = 'TEST prec total'
        c = 30
        for i in range(acc.shape[0]):
            sheet.cell(row=c, column=1).value = 'TEST ACC ID' + str(i)
            c += 1
            sheet.cell(row=c, column=1).value = 'TEST PREC ID' + str(i)
            c += 1
            sheet.cell(row=c, column=1).value = 'TEST F1 ID' + str(i)
            c += 1
    # Update Excel file in the next empty col
    col = len([cell for cell in sheet[1] if cell.value is not None]) + 1
    sheet.cell(row=1, column=col).value = data[0]
    sheet.cell(row=2, column=col).value = data[1]
    sheet.cell(row=3, column=col).value = data[2]
    sheet.cell(row=4, column=col).value = data[3]
    sheet.cell(row=5, column=col).value = data[4]
    txt = ''
    for i in data[5]:
        txt += str(i) + '-'
    sheet.cell(row=6, column=col).value = txt[:-1]
    sheet.cell(row=7, column=col).value = data[6]
    sheet.cell(row=8, column=col).value = data[7]
    sheet.cell(row=9, column=col).value = data[8]
    sheet.cell(row=10, column=col).value = data[9]
    sheet.cell(row=11, column=col).value = data[10]
    sheet.cell(row=12, column=col).value = train[0]
    sheet.cell(row=13, column=col).value = val[0]
    sheet.cell(row=14, column=col).value = test[0]
    sheet.cell(row=15, column=col).value = train[1]
    sheet.cell(row=16, column=col).value = val[1]
    sheet.cell(row=17, column=col).value = test[1]
    sheet.cell(row=18, column=col).value = train[2]
    sheet.cell(row=19, column=col).value = val[2]
    sheet.cell(row=20, column=col).value = test[2]
    sheet.cell(row=21, column=col).value = train[3]
    sheet.cell(row=22, column=col).value = val[3]
    sheet.cell(row=23, column=col).value = test[3]
    sheet.cell(row=24, column=col).value = train[4]
    sheet.cell(row=25, column=col).value = val[4]
    sheet.cell(row=26, column=col).value = test[4]
    sheet.cell(row=27, column=col).value = train[5]
    sheet.cell(row=28, column=col).value = val[5]
    sheet.cell(row=29, column=col).value = test[5]
    c = 30
    for i in range(acc.shape[0]):
        sheet.cell(row=c, column=col).value = round(acc[i], 1)
        c += 1
        sheet.cell(row=c, column=col).value = round(prec[i], 1)
        c += 1
        sheet.cell(row=c, column=col).value = round(f1[i], 1)
        c += 1
    excel_file.save(file + '.xlsx')
